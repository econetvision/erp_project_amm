import os
import base64
import uuid
import random
import re
from datetime import date, datetime
from io import BytesIO
from fastapi import APIRouter, Depends, HTTPException, Query, UploadFile, File
from fastapi.responses import StreamingResponse
from pydantic import BaseModel, ValidationError
from sqlalchemy.orm import Session
from sqlalchemy import or_
from database import get_db
from models.user import User
from schemas.employee import (
    EmployeeCreate, EmployeeUpdate, EmployeeResponse,
    EmployeeCreateResponse, WorkLocationUpdateSchema, EmployeeCodeUpdateSchema
)
from services.face_service import get_face_encoding
from services.license_service import enforce_seat_limit, SEAT_LIMIT
from services import storage
from auth.dependencies import (
    require_admin_or_supervisor, require_admin, get_current_user,
    hash_password, assert_tenant, tenant_scope,
)
from config.settings import settings

router = APIRouter()


# Defaults for the employee-code pattern; overridable per company via
# Company.employee_id_config (set in Company Settings → Employee ID).
DEFAULT_EMPLOYEE_ID_CONFIG = {
    "prefix": "",          # company segment; empty → derived from company name
    "include_site": True,  # append a segment from the work-location name
    "separator": "-",
    "seq_digits": 3,       # zero-padding of the running number
    "seq_start": 1,        # first number when the prefix has no employees yet
}


def _employee_id_config(company) -> dict:
    cfg = dict(DEFAULT_EMPLOYEE_ID_CONFIG)
    if company is not None and isinstance(company.employee_id_config, dict):
        cfg.update({k: v for k, v in company.employee_id_config.items() if v is not None})
    # Sanitize: separator limited to safe characters, numeric fields to sane ranges
    if cfg.get("separator") not in ("-", "/", "_", ""):
        cfg["separator"] = "-"
    try:
        cfg["seq_digits"] = min(max(int(cfg.get("seq_digits", 3)), 1), 8)
    except (TypeError, ValueError):
        cfg["seq_digits"] = 3
    try:
        cfg["seq_start"] = max(int(cfg.get("seq_start", 1)), 1)
    except (TypeError, ValueError):
        cfg["seq_start"] = 1
    return cfg


def _generate_employee_code(db: Session, company_id: int | None, work_location_name: str | None = None) -> str:
    """Generate the next employee code following the company's configured pattern.

    Pattern: PREFIX[<sep>SITE]<sep>SEQ — e.g. ABC-HQ-001, XYZ/PLANT1/0042.
    The prefix, site segment, separator, padding and starting number come from
    Company.employee_id_config; without a config it falls back to the legacy
    COMPANY-SITE-NNN derived from the company name.
    """
    company = None
    if company_id:
        from models.company import Company
        company = db.query(Company).filter(Company.id == company_id).first()

    cfg = _employee_id_config(company)
    sep = cfg["separator"]

    prefix = re.sub(r"[^A-Za-z0-9]", "", str(cfg.get("prefix") or "")).upper()
    if not prefix:
        # Legacy behaviour: first 5 letters of the company name
        letters = re.sub(r"[^A-Za-z]", "", company.name).upper() if company else ""
        prefix = letters[:5] or "EMP"

    parts = [prefix]
    if cfg["include_site"]:
        site_prefix = "HQ"
        if work_location_name:
            site_letters = re.sub(r"[^A-Za-z0-9]", "", work_location_name).upper()
            site_prefix = site_letters[:5] or "HQ"
        parts.append(site_prefix)

    base = sep.join(parts) + sep

    # Next sequence = max numeric suffix among existing codes with this base
    existing_codes = db.query(User.employee_code).filter(
        User.employee_code.like(f"{base}%")
    ).all()
    max_id = 0
    for (code,) in existing_codes:
        if code:
            m = re.search(r"(\d+)$", code[len(base):])
            if m:
                max_id = max(max_id, int(m.group(1)))

    next_id = max_id + 1 if max_id else cfg["seq_start"]
    return f"{base}{next_id:0{cfg['seq_digits']}d}"


# Fields a caller may mass-edit via PUT/PATCH /{employee_id}. Derived from the
# EmployeeUpdate schema MINUS a hard denylist of privilege/identity fields that
# must never be set through the generic update path (role/company escalation,
# credential or activation tampering, employee-code collisions).
_EMPLOYEE_UPDATE_DENYLIST = {"role", "company_id", "username", "password_hash", "is_active", "employee_code"}
EMPLOYEE_EDITABLE_FIELDS = set(EmployeeUpdate.model_fields) - _EMPLOYEE_UPDATE_DENYLIST


class FaceRegisterRequest(BaseModel):
    image: str  # base64-encoded image


@router.get("")
def list_employees(
    page: int = Query(1, ge=1),
    per_page: int = Query(10, ge=1, le=100),
    search: str = Query("", alias="q"),
    all: bool = Query(False),
    db: Session = Depends(get_db),
    current_user: User = Depends(require_admin_or_supervisor),
):
    # Supervisors can only see workers, admins/masters can see both workers and supervisors
    if current_user.role == "supervisor":
        q = db.query(User).filter(User.role == "worker")
    else:
        q = db.query(User).filter(User.role.in_(["worker", "supervisor"]))
    # Tenant isolation: non-master callers only see their own company's employees.
    q = tenant_scope(q, User.company_id, current_user)
    if search:
        q = q.filter(
            or_(
                User.name.ilike(f"%{search}%"),
                User.aadhar_number.ilike(f"%{search}%"),
                User.bank_account_number.ilike(f"%{search}%"),
            )
        )
    if all:
        items = q.order_by(User.id).all()
        return {
            "items": [EmployeeResponse.model_validate(e) for e in items],
            "total": len(items),
            "page": 1,
            "per_page": len(items),
            "pages": 1,
        }
    total = q.count()
    items = q.order_by(User.id).offset((page - 1) * per_page).limit(per_page).all()
    return {
        "items": [EmployeeResponse.model_validate(e) for e in items],
        "total": total,
        "page": page,
        "per_page": per_page,
        "pages": max(1, -(-total // per_page)),
    }


def _persist_employee(db: Session, payload: EmployeeCreate, current_user: User) -> User:
    """Validate and insert one employee. Shared by the create endpoint and bulk import."""
    if not payload.name or not payload.name.strip():
        raise HTTPException(status_code=422, detail="Employee name is required")

    # Check Aadhar uniqueness
    existing = db.query(User).filter(User.aadhar_number == payload.aadhar_number).first()
    if existing:
        raise HTTPException(status_code=400, detail="Aadhar number already registered")

    data = payload.model_dump()

    # Admin-defined username and password (required)
    username = data.pop("username")
    password = data.pop("password")
    role = data.pop("role", None) or "worker"
    # Tenant isolation: only a master may target another company; every other
    # caller is pinned to their own company_id regardless of any client-supplied value.
    supplied_company_id = data.pop("company_id", None)
    if current_user.role == "master":
        company_id = supplied_company_id or current_user.company_id
    else:
        company_id = current_user.company_id

    # Check username uniqueness
    if db.query(User).filter(User.username == username).first():
        raise HTTPException(status_code=400, detail="Username already exists. Please choose a different username.")

    # Supervisors can only create workers
    if current_user.role == "supervisor" and role != "worker":
        raise HTTPException(status_code=403, detail="Supervisors can only create workers")

    # Every new employee consumes a license seat — enforce the company's seat
    # limit (no-op when license enforcement is bypassed via settings flags).
    if company_id is not None:
        enforce_seat_limit(db, company_id)

    # Generate employee code following the company's configured pattern
    work_location = data.get("work_location_name")
    employee_code = _generate_employee_code(db, company_id, work_location)

    emp = User(
        username=username,
        password_hash=hash_password(password),
        role=role,
        company_id=company_id,
        employee_code=employee_code,
        onboarding_complete=False,
        **data,
    )
    db.add(emp)
    db.commit()
    db.refresh(emp)
    return emp


@router.post("", status_code=201)
def create_employee(payload: EmployeeCreate, db: Session = Depends(get_db), current_user: User = Depends(require_admin_or_supervisor)):
    """Create a new employee with admin-defined login credentials.

    Admin must provide username and password for the employee.
    Employee ID is auto-generated following the company's configured pattern.
    """
    emp = _persist_employee(db, payload, current_user)

    return {
        "id": emp.id,
        "employee_code": emp.employee_code,
        "username": emp.username,
        "role": emp.role,
        "company_id": emp.company_id,
        "name": emp.name,
        "aadhar_number": emp.aadhar_number,
        "onboarding_complete": emp.onboarding_complete,
        "created_at": emp.created_at,
    }


# ── Bulk import / export ──────────────────────────────────────────────────────
# NOTE: these static routes must stay declared before GET /{employee_id}.

# (column key, required, description, example) — drives both the sample template
# and the import parser. Keys match EmployeeCreate fields.
IMPORT_COLUMNS = [
    ("username",            True,  "Login username (letters, numbers, underscore only)", "john_d"),
    ("password",            True,  "Initial login password (min 6 characters)",          "Secret123"),
    ("name",                True,  "Full name",                                          "John Doe"),
    ("address",             True,  "Residential address",                                "12 MG Road, Pune"),
    ("aadhar_number",       True,  "Exactly 12 digits",                                  "234567890123"),
    ("bank_account_number", True,  "8-18 digits",                                        "12345678901"),
    ("monthly_salary",      False, "Monthly salary, number >= 0 (falls back to hourly rate if blank)", "26000.00"),
    ("hourly_rate",         True,  "Hourly wage (used for overtime), number >= 0",       "150.50"),
    ("gender",              False, "male / female / other",                              "male"),
    ("date_of_birth",       False, "YYYY-MM-DD",                                         "1995-04-23"),
    ("blood_group",         False, "e.g. O+, AB-",                                       "O+"),
    ("marital_status",      False, "single / married / divorced / widowed",              "single"),
    ("emergency_name",      False, "Emergency contact person",                           "Jane Doe"),
    ("emergency_contact",   False, "Emergency contact phone",                            "9876543210"),
    ("phone",               False, "Phone number",                                       "9876501234"),
    ("email",               False, "Email address",                                      "john@example.com"),
    ("ifsc_code",           False, "11 chars: 4 letters + 0 + 6 alphanumeric",           "HDFC0001234"),
    ("shift",               False, "SHIFT_A or SHIFT_B (default: SHIFT_A)",              "SHIFT_A"),
    ("work_location_name",  False, "Assigned work location name",                        "HQ"),
    ("work_latitude",       False, "Latitude (decimal degrees)",                         "18.5204"),
    ("work_longitude",      False, "Longitude (decimal degrees)",                        "73.8567"),
    ("attendance_radius_m", False, "Geofence radius in metres (default: 50)",            "100"),
]

EXPORT_COLUMNS = [
    "employee_code", "username", "role", "name", "gender", "date_of_birth",
    "blood_group", "marital_status", "emergency_name", "emergency_contact",
    "phone", "email", "address", "aadhar_number", "bank_account_number",
    "ifsc_code", "bank_name", "kyc_status", "monthly_salary", "hourly_rate", "shift",
    "work_location_name", "work_latitude", "work_longitude",
    "attendance_radius_m", "onboarding_complete", "created_at",
]

MAX_IMPORT_ROWS = 1000
MAX_IMPORT_BYTES = 5 * 1024 * 1024  # 5 MB


def _xlsx_response(workbook, filename: str) -> StreamingResponse:
    buf = BytesIO()
    workbook.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@router.get("/import/template")
def download_import_template(current_user: User = Depends(require_admin)):
    """Sample Excel template for bulk employee import, with an instructions sheet."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill
    from openpyxl.utils import get_column_letter

    wb = Workbook()

    ws = wb.active
    ws.title = "Employees"
    header_font = Font(bold=True, color="FFFFFF")
    required_fill = PatternFill("solid", fgColor="0D6EFD")
    optional_fill = PatternFill("solid", fgColor="6C757D")
    for col, (key, required, _desc, example) in enumerate(IMPORT_COLUMNS, start=1):
        cell = ws.cell(row=1, column=col, value=f"{key}{'*' if required else ''}")
        cell.font = header_font
        cell.fill = required_fill if required else optional_fill
        ws.cell(row=2, column=col, value=example)
        ws.column_dimensions[get_column_letter(col)].width = max(len(key) + 3, 14)
    # Text format for identifier columns so Excel doesn't mangle long numbers
    for col, (key, _r, _d, _e) in enumerate(IMPORT_COLUMNS, start=1):
        if key in ("aadhar_number", "bank_account_number", "phone", "emergency_contact"):
            for row in range(2, 200):
                ws.cell(row=row, column=col).number_format = "@"

    info = wb.create_sheet("Instructions")
    info.append(["Bulk Employee Import — Instructions"])
    info["A1"].font = Font(bold=True, size=14)
    info.append([])
    info.append(["1. Fill one employee per row in the 'Employees' sheet (row 2 is an example — replace it)."])
    info.append(["2. Columns marked with * are required. Do not rename or reorder the header row."])
    info.append(["3. Dates must be YYYY-MM-DD. Keep Aadhar/account numbers as text to avoid losing digits."])
    info.append(["4. The Employee ID is generated automatically from your company's configured pattern."])
    info.append(["5. Rows that fail validation are skipped and reported — the rest are still imported."])
    info.append(["6. All imported employees are created as workers. Create supervisor/admin accounts individually from User Management."])
    info.append([])
    info.append(["Column", "Required", "Description", "Example"])
    for c in ("A", "B", "C", "D"):
        info[f"{c}9"].font = Font(bold=True)
    for key, required, desc, example in IMPORT_COLUMNS:
        info.append([key, "Yes" if required else "No", desc, example])
    info.column_dimensions["A"].width = 24
    info.column_dimensions["B"].width = 10
    info.column_dimensions["C"].width = 52
    info.column_dimensions["D"].width = 22

    return _xlsx_response(wb, "employee_import_template.xlsx")


@router.get("/export")
def export_employees(
    search: str = Query("", alias="q"),
    db: Session = Depends(get_db),
    current_user: User = Depends(require_admin),
):
    """Export all employees (workers + supervisors) to an Excel file."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill
    from openpyxl.utils import get_column_letter

    q = db.query(User).filter(User.role.in_(["worker", "supervisor"]))
    # Tenant isolation: the export includes Aadhaar/bank details — never leak
    # another company's employees to a non-master admin.
    q = tenant_scope(q, User.company_id, current_user)
    if search:
        q = q.filter(
            or_(
                User.name.ilike(f"%{search}%"),
                User.aadhar_number.ilike(f"%{search}%"),
                User.bank_account_number.ilike(f"%{search}%"),
            )
        )
    employees = q.order_by(User.id).all()

    wb = Workbook()
    ws = wb.active
    ws.title = "Employees"
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="0D6EFD")
    for col, key in enumerate(EXPORT_COLUMNS, start=1):
        cell = ws.cell(row=1, column=col, value=key)
        cell.font = header_font
        cell.fill = header_fill
        ws.column_dimensions[get_column_letter(col)].width = max(len(key) + 3, 14)

    for row, emp in enumerate(employees, start=2):
        for col, key in enumerate(EXPORT_COLUMNS, start=1):
            value = getattr(emp, key, None)
            if isinstance(value, datetime):
                value = value.strftime("%Y-%m-%d %H:%M:%S")
            elif isinstance(value, date):
                value = value.isoformat()
            elif value is not None and key in ("hourly_rate", "monthly_salary"):
                value = float(value)
            cell = ws.cell(row=row, column=col, value=value)
            if key in ("aadhar_number", "bank_account_number", "phone", "emergency_contact"):
                cell.number_format = "@"

    return _xlsx_response(wb, f"employees_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx")


def _cell_to_str(value):
    """Excel cells often hold numbers where we need strings (Aadhar, phone…)."""
    if value is None:
        return None
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    s = str(value).strip()
    return s or None


def _parse_import_row(raw: dict) -> dict:
    """Coerce one spreadsheet row into EmployeeCreate keyword arguments."""
    data = {}
    for key, _required, _desc, _example in IMPORT_COLUMNS:
        value = raw.get(key)
        if value is None or (isinstance(value, str) and not value.strip()):
            continue
        if key in ("hourly_rate", "monthly_salary", "work_latitude", "work_longitude", "attendance_radius_m"):
            data[key] = float(value)
        elif key == "date_of_birth":
            if isinstance(value, datetime):
                data[key] = value.date()
            elif isinstance(value, date):
                data[key] = value
            else:
                data[key] = str(value).strip()[:10]  # pydantic parses YYYY-MM-DD
        else:
            data[key] = _cell_to_str(value)

    if data.get("shift"):
        shift = data["shift"].upper().replace(" ", "_")
        data["shift"] = {"A": "SHIFT_A", "B": "SHIFT_B"}.get(shift, shift)
    for key in ("gender", "marital_status"):
        if data.get(key):
            data[key] = data[key].lower()
    if data.get("ifsc_code"):
        data["ifsc_code"] = data["ifsc_code"].upper()
    return data


@router.post("/import")
async def import_employees(
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    current_user: User = Depends(require_admin),
):
    """Bulk-create employees from an .xlsx file (see /import/template for the format).

    Each row is validated and inserted independently: valid rows are created,
    invalid rows are reported with their row number and reason.
    """
    from openpyxl import load_workbook

    # Fail fast when the company's license is missing/suspended/expired or the
    # seat limit is already reached, before parsing the file at all.
    if current_user.company_id is not None:
        enforce_seat_limit(db, current_user.company_id)

    if file.filename and not file.filename.lower().endswith((".xlsx", ".xlsm")):
        raise HTTPException(status_code=400, detail="Please upload an Excel (.xlsx) file — download the sample template first")

    content = await file.read()
    if len(content) > MAX_IMPORT_BYTES:
        raise HTTPException(status_code=400, detail="File too large (max 5 MB)")

    try:
        wb = load_workbook(BytesIO(content), data_only=True, read_only=True)
    except Exception:
        raise HTTPException(status_code=400, detail="Could not read the file as an Excel workbook")

    ws = wb["Employees"] if "Employees" in wb.sheetnames else wb.active
    rows = ws.iter_rows(values_only=True)
    try:
        header_row = next(rows)
    except StopIteration:
        raise HTTPException(status_code=400, detail="The spreadsheet is empty")

    # Normalize headers: "Aadhar Number*" -> "aadhar_number"
    headers = [
        re.sub(r"\s+", "_", str(h).strip().strip("*").lower()) if h else None
        for h in header_row
    ]
    known_keys = {key for key, *_ in IMPORT_COLUMNS}
    if not known_keys.intersection(h for h in headers if h):
        raise HTTPException(status_code=400, detail="No recognised columns found — download the sample template and keep its header row")

    created, errors = [], []
    row_count = 0
    for idx, row in enumerate(rows, start=2):
        if row is None or all(v is None or str(v).strip() == "" for v in row):
            continue
        row_count += 1
        if row_count > MAX_IMPORT_ROWS:
            errors.append({"row": idx, "error": f"Import limited to {MAX_IMPORT_ROWS} rows per file — remaining rows skipped"})
            break

        raw = {key: value for key, value in zip(headers, row) if key}

        # Bulk import creates workers only; reject rows that ask for another
        # role (e.g. files based on an older template) instead of silently
        # downgrading them.
        role_value = _cell_to_str(raw.get("role"))
        if role_value and role_value.lower() != "worker":
            errors.append({"row": idx, "error": "Bulk import can only create workers — remove the role column or set it to 'worker'"})
            continue

        try:
            payload = EmployeeCreate(**_parse_import_row(raw))
            emp = _persist_employee(db, payload, current_user)
            created.append({"row": idx, "employee_code": emp.employee_code, "name": emp.name})
        except ValidationError as ve:
            errors.append({
                "row": idx,
                "error": "; ".join(
                    f"{'.'.join(str(loc) for loc in err['loc'])}: {err['msg']}" for err in ve.errors()
                ),
            })
        except HTTPException as he:
            db.rollback()
            if he.detail == SEAT_LIMIT:
                errors.append({"row": idx, "error": "License seat limit reached — this and all remaining rows were skipped"})
                break
            errors.append({"row": idx, "error": str(he.detail)})
        except Exception as e:
            db.rollback()
            errors.append({"row": idx, "error": str(e)})

    return {
        "total_rows": row_count,
        "created": len(created),
        "failed": len(errors),
        "created_employees": created,
        "errors": errors,
    }


@router.get("/{employee_id}", response_model=EmployeeResponse)
def get_employee(employee_id: int, db: Session = Depends(get_db), current_user: User = Depends(require_admin_or_supervisor)):
    emp = db.query(User).filter(User.id == employee_id).first()
    if not emp:
        raise HTTPException(status_code=404, detail="Employee not found")
    assert_tenant(current_user, emp.company_id)

    # Supervisors can only view workers (not other supervisors)
    if current_user.role == "supervisor" and emp.role != "worker":
        raise HTTPException(status_code=403, detail="Supervisors can only view workers")

    return emp


@router.put("/{employee_id}", response_model=EmployeeResponse)
def update_employee(employee_id: int, payload: EmployeeUpdate, db: Session = Depends(get_db), current_user: User = Depends(require_admin)):
    """Full update of employee - Admin only.

    Supervisors should use PATCH /employees/{id}/work-location for work location updates.
    Note: Username and password cannot be changed via this endpoint.
    """
    emp = db.query(User).filter(User.id == employee_id).first()
    if not emp:
        raise HTTPException(status_code=404, detail="Employee not found")
    assert_tenant(current_user, emp.company_id)

    # Update only provided, allow-listed fields (never role/company/credentials).
    for key, value in payload.model_dump(exclude_none=True).items():
        if key in EMPLOYEE_EDITABLE_FIELDS:
            setattr(emp, key, value)
    db.commit()
    db.refresh(emp)
    return emp


@router.patch("/{employee_id}", response_model=EmployeeResponse)
def partial_update_employee(employee_id: int, payload: EmployeeUpdate, db: Session = Depends(get_db), current_user: User = Depends(require_admin)):
    """Partial update of employee - Admin only.

    Supervisors should use PATCH /employees/{id}/work-location for work location updates.
    """
    emp = db.query(User).filter(User.id == employee_id).first()
    if not emp:
        raise HTTPException(status_code=404, detail="Employee not found")
    assert_tenant(current_user, emp.company_id)

    # Update only provided, allow-listed fields (never role/company/credentials).
    for key, value in payload.model_dump(exclude_none=True).items():
        if key in EMPLOYEE_EDITABLE_FIELDS:
            setattr(emp, key, value)
    db.commit()
    db.refresh(emp)
    return emp


@router.patch("/{employee_id}/work-location", response_model=EmployeeResponse)
def update_employee_work_location(
    employee_id: int,
    payload: WorkLocationUpdateSchema,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_admin_or_supervisor)
):
    """Update employee work location - Available to Admin and Supervisor.

    Supervisors can only update work location for workers, not other supervisors.
    """
    emp = db.query(User).filter(User.id == employee_id).first()
    if not emp:
        raise HTTPException(status_code=404, detail="Employee not found")
    assert_tenant(current_user, emp.company_id)

    # Supervisors can only edit workers
    if current_user.role == "supervisor" and emp.role != "worker":
        raise HTTPException(status_code=403, detail="Supervisors can only edit work location for workers")

    for key, value in payload.model_dump(exclude_none=True).items():
        setattr(emp, key, value)
    db.commit()
    db.refresh(emp)
    return emp


@router.patch("/{employee_id}/employee-code", response_model=EmployeeResponse)
def update_employee_code(
    employee_id: int,
    payload: EmployeeCodeUpdateSchema,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_admin)
):
    """Update employee code - Admin only.

    Employee code format: COMPANY-SITE-ID (e.g., ABC-HQ-001)
    """
    emp = db.query(User).filter(User.id == employee_id).first()
    if not emp:
        raise HTTPException(status_code=404, detail="Employee not found")
    assert_tenant(current_user, emp.company_id)

    # Check if the new code is already in use
    existing = db.query(User).filter(
        User.employee_code == payload.employee_code,
        User.id != employee_id
    ).first()
    if existing:
        raise HTTPException(status_code=400, detail="Employee code already in use")

    emp.employee_code = payload.employee_code
    db.commit()
    db.refresh(emp)
    return emp


@router.delete("/{employee_id}", status_code=204)
def delete_employee(employee_id: int, db: Session = Depends(get_db), current_user: User = Depends(require_admin)):
    """Delete employee - Admin only."""
    emp = db.query(User).filter(User.id == employee_id).first()
    if not emp:
        raise HTTPException(status_code=404, detail="Employee not found")
    assert_tenant(current_user, emp.company_id)
    if emp.id == current_user.id:
        raise HTTPException(status_code=400, detail="Cannot delete your own account")
    db.delete(emp)
    db.commit()


@router.post("/{employee_id}/face", response_model=EmployeeResponse)
def register_face(employee_id: int, payload: FaceRegisterRequest, db: Session = Depends(get_db), current_user: User = Depends(require_admin_or_supervisor)):
    """Register employee face for biometric attendance.

    This is a critical step in employee onboarding. Once face is registered,
    the employee can use face scan for attendance and login.
    """
    emp = db.query(User).filter(User.id == employee_id).first()
    if not emp:
        raise HTTPException(status_code=404, detail="Employee not found")
    assert_tenant(current_user, emp.company_id)

    # Supervisors can only register faces for workers
    if current_user.role == "supervisor" and emp.role != "worker":
        raise HTTPException(status_code=403, detail="Supervisors can only register faces for workers")

    emp.face_encoding = get_face_encoding(payload.image)
    # Save photo as file instead of base64
    try:
        img_data = payload.image
        if "," in img_data:
            img_data = img_data.split(",", 1)[1]
        img_bytes = base64.b64decode(img_data)
    except Exception:
        raise HTTPException(status_code=400, detail="Invalid image data")
    if len(img_bytes) > 5 * 1024 * 1024:
        raise HTTPException(status_code=400, detail="Image must be under 5 MB")
    filename = f"{employee_id}_{uuid.uuid4().hex[:8]}.jpg"
    emp.photo = storage.save_image("employees", filename, img_bytes, "image/jpeg")

    # Mark onboarding as complete once face is registered
    emp.onboarding_complete = True

    db.commit()
    db.refresh(emp)
    return emp


@router.get("/{employee_id}/ifsc-lookup")
async def ifsc_lookup(employee_id: int, ifsc: str, db: Session = Depends(get_db), current_user: User = Depends(require_admin_or_supervisor)):
    from services.ifsc_service import lookup_ifsc
    emp = db.query(User).filter(User.id == employee_id).first()
    if not emp:
        raise HTTPException(status_code=404, detail="Employee not found")
    assert_tenant(current_user, emp.company_id)
    return await lookup_ifsc(ifsc, db)


@router.post("/{employee_id}/verify-bank", response_model=EmployeeResponse)
async def verify_bank_account_endpoint(
    employee_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_admin_or_supervisor),
):
    from services.kyc_service import verify_bank_account
    import os
    emp = db.query(User).filter(User.id == employee_id).first()
    if not emp:
        raise HTTPException(status_code=404, detail="Employee not found")
    assert_tenant(current_user, emp.company_id)
    if not emp.ifsc_code:
        raise HTTPException(status_code=400, detail="IFSC code is required for bank verification")
    
    result = await verify_bank_account(
        bank_account=emp.bank_account_number,
        ifsc_code=emp.ifsc_code,
        account_holder_name=emp.name,
        provider=settings.kyc_provider,
        api_key=settings.kyc_api_key or None,
        api_secret=settings.kyc_api_secret or None,
    )
    emp.kyc_status = result["status"]
    emp.kyc_verified_name = result.get("registered_name")
    db.commit()
    db.refresh(emp)
    return emp


# ── Twilio Phone/Email Verification ──────────────────────────────────────────

class OTPRequest(BaseModel):
    code: str


@router.post("/{employee_id}/send-phone-otp")
def send_phone_verification(employee_id: int, db: Session = Depends(get_db), current_user: User = Depends(require_admin_or_supervisor)):
    """Send OTP to employee's phone number via Twilio."""
    from services.twilio_service import send_phone_otp
    emp = db.query(User).filter(User.id == employee_id).first()
    if not emp:
        raise HTTPException(status_code=404, detail="Employee not found")
    assert_tenant(current_user, emp.company_id)
    if not emp.phone:
        raise HTTPException(status_code=400, detail="Employee phone number is not set")
    result = send_phone_otp(emp.phone, db=db, company_id=emp.company_id)
    if result.get("status") == "error":
        raise HTTPException(status_code=502, detail=result.get("message", "Failed to send OTP"))
    return {"detail": "OTP sent to phone", "status": result["status"]}


@router.post("/{employee_id}/verify-phone", response_model=EmployeeResponse)
def verify_phone(employee_id: int, payload: OTPRequest, db: Session = Depends(get_db), current_user: User = Depends(require_admin_or_supervisor)):
    """Verify phone OTP and mark phone as verified."""
    from services.twilio_service import verify_phone_otp
    emp = db.query(User).filter(User.id == employee_id).first()
    if not emp:
        raise HTTPException(status_code=404, detail="Employee not found")
    assert_tenant(current_user, emp.company_id)
    if not emp.phone:
        raise HTTPException(status_code=400, detail="Employee phone number is not set")
    result = verify_phone_otp(emp.phone, payload.code, db=db, company_id=emp.company_id)
    if not result.get("valid"):
        raise HTTPException(status_code=400, detail="Invalid or expired OTP")
    emp.phone_verified = "Y"
    db.commit()
    db.refresh(emp)
    return emp


@router.post("/{employee_id}/send-email-otp")
def send_email_verification(employee_id: int, db: Session = Depends(get_db), current_user: User = Depends(require_admin_or_supervisor)):
    """Send OTP to employee's email via Twilio."""
    from services.twilio_service import send_email_otp
    emp = db.query(User).filter(User.id == employee_id).first()
    if not emp:
        raise HTTPException(status_code=404, detail="Employee not found")
    assert_tenant(current_user, emp.company_id)
    if not emp.email:
        raise HTTPException(status_code=400, detail="Employee email is not set")
    result = send_email_otp(emp.email, db=db, company_id=emp.company_id)
    if result.get("status") == "error":
        raise HTTPException(status_code=502, detail=result.get("message", "Failed to send OTP"))
    return {"detail": "OTP sent to email", "status": result["status"]}


@router.post("/{employee_id}/verify-email", response_model=EmployeeResponse)
def verify_email(employee_id: int, payload: OTPRequest, db: Session = Depends(get_db), current_user: User = Depends(require_admin_or_supervisor)):
    """Verify email OTP and mark email as verified."""
    from services.twilio_service import verify_email_otp
    emp = db.query(User).filter(User.id == employee_id).first()
    if not emp:
        raise HTTPException(status_code=404, detail="Employee not found")
    assert_tenant(current_user, emp.company_id)
    if not emp.email:
        raise HTTPException(status_code=400, detail="Employee email is not set")
    result = verify_email_otp(emp.email, payload.code, db=db, company_id=emp.company_id)
    if not result.get("valid"):
        raise HTTPException(status_code=400, detail="Invalid or expired OTP")
    emp.email_verified = "Y"
    db.commit()
    db.refresh(emp)
    return emp


# ── Work Location Assignment ─────────────────────────────────────────────────

class WorkLocationAssignment(BaseModel):
    work_location_name: str
    work_latitude: float
    work_longitude: float
    attendance_radius_m: float = 50.0


@router.put("/{employee_id}/work-location", response_model=EmployeeResponse)
def assign_work_location(employee_id: int, payload: WorkLocationAssignment, db: Session = Depends(get_db), current_user: User = Depends(require_admin_or_supervisor)):
    """Assign or update work location for an User."""
    emp = db.query(User).filter(User.id == employee_id).first()
    if not emp:
        raise HTTPException(status_code=404, detail="Employee not found")
    assert_tenant(current_user, emp.company_id)
    emp.work_location_name = payload.work_location_name
    emp.work_latitude = payload.work_latitude
    emp.work_longitude = payload.work_longitude
    emp.attendance_radius_m = payload.attendance_radius_m
    db.commit()
    db.refresh(emp)
    return emp
